# -*- coding: utf-8 -*-
"""
Exitus - Export Service (EXITUS-EXPORT-001)
Exportação de dados para CSV, Excel, JSON e PDF.

Entidades exportáveis:
- transacoes: compras, vendas, dividendos, etc.
- proventos:  dividendos, JCP, aluguéis
- posicoes:   posição consolidada do portfólio

Filtros disponíveis (todos opcionais):
- data_inicio / data_fim  (YYYY-MM-DD)
- ativo_id
- corretora_id
- tipo  (depende da entidade)

Formatos: csv | excel | json | pdf
"""

import csv
import io
import json
import logging
from datetime import datetime, date
from decimal import Decimal

from app.database import db
from app.models.transacao import Transacao, TipoTransacao
from app.models.ativo import Ativo
from app.models.corretora import Corretora
from app.utils.exceptions import BusinessRuleError

logger = logging.getLogger(__name__)

FORMATOS_VALIDOS = {'csv', 'excel', 'json', 'pdf'}
LIMITE_REGISTROS = 10_000


# ---------------------------------------------------------------------------
# Helpers de formatação
# ---------------------------------------------------------------------------

def _fmt_data(dt) -> str:
    if dt is None:
        return ''
    if hasattr(dt, 'strftime'):
        return dt.strftime('%d/%m/%Y')
    return str(dt)


def _fmt_decimal(v) -> str:
    if v is None:
        return ''
    return f'{float(v):.2f}'.replace('.', ',')


def _fmt_decimal_f(v) -> float:
    if v is None:
        return 0.0
    return round(float(v), 2)


def _parse_date(s: str, campo: str) -> date:
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except Exception:
        raise BusinessRuleError(f"Formato inválido para '{campo}': use YYYY-MM-DD.")


# ---------------------------------------------------------------------------
# Queries por entidade
# ---------------------------------------------------------------------------

def _query_transacoes(usuario_id: str, filtros: dict) -> list[dict]:
    q = (
        Transacao.query
        .filter(Transacao.usuario_id == usuario_id)
        .join(Ativo, Transacao.ativo_id == Ativo.id)
        .join(Corretora, Transacao.corretora_id == Corretora.id)
        .add_columns(Ativo.ticker, Ativo.nome.label('ativo_nome'), Corretora.nome.label('corretora_nome'))
        .order_by(Transacao.data_transacao.desc())
    )

    if filtros.get('data_inicio'):
        q = q.filter(Transacao.data_transacao >= filtros['data_inicio'])
    if filtros.get('data_fim'):
        q = q.filter(Transacao.data_transacao <= filtros['data_fim'])
    if filtros.get('ativo_id'):
        q = q.filter(Transacao.ativo_id == filtros['ativo_id'])
    if filtros.get('corretora_id'):
        q = q.filter(Transacao.corretora_id == filtros['corretora_id'])
    if filtros.get('tipo'):
        try:
            q = q.filter(Transacao.tipo == TipoTransacao(filtros['tipo'].lower()))
        except ValueError:
            raise BusinessRuleError(f"Tipo inválido: '{filtros['tipo']}'.")

    rows = q.limit(LIMITE_REGISTROS).all()

    resultado = []
    for t, ticker, ativo_nome, corretora_nome in rows:
        resultado.append({
            'id':               str(t.id),
            'data':             _fmt_data(t.data_transacao),
            'tipo':             t.tipo.value,
            'ticker':           ticker,
            'ativo':            ativo_nome,
            'corretora':        corretora_nome,
            'quantidade':       _fmt_decimal_f(t.quantidade),
            'preco_unitario':   _fmt_decimal_f(t.preco_unitario),
            'valor_total':      _fmt_decimal_f(t.valor_total),
            'custos_totais':    _fmt_decimal_f(t.custos_totais),
            'valor_liquido':    _fmt_decimal_f(t.valor_liquido),
            'observacoes':      t.observacoes or '',
        })
    return resultado


def _query_proventos(usuario_id: str, filtros: dict) -> list[dict]:
    from app.models.provento import Provento, TipoProvento
    from app.models.ativo import Ativo as AtivoModel

    # Provento não tem usuario_id direto — filtra via ativo_id das transações do usuário
    from app.models.transacao import Transacao
    ativos_do_usuario = (
        db.session.query(Transacao.ativo_id)
        .filter(Transacao.usuario_id == usuario_id)
        .distinct()
        .subquery()
    )

    q = (
        Provento.query
        .filter(Provento.ativo_id.in_(ativos_do_usuario))
        .join(AtivoModel, Provento.ativo_id == AtivoModel.id)
        .add_columns(AtivoModel.ticker, AtivoModel.nome.label('ativo_nome'))
        .order_by(Provento.data_pagamento.desc())
    )

    if filtros.get('data_inicio'):
        q = q.filter(Provento.data_pagamento >= filtros['data_inicio'])
    if filtros.get('data_fim'):
        q = q.filter(Provento.data_pagamento <= filtros['data_fim'])
    if filtros.get('ativo_id'):
        q = q.filter(Provento.ativo_id == filtros['ativo_id'])
    if filtros.get('tipo'):
        try:
            q = q.filter(Provento.tipo_provento == TipoProvento(filtros['tipo'].lower()))
        except ValueError:
            raise BusinessRuleError(f"Tipo de provento inválido: '{filtros['tipo']}'.")

    rows = q.limit(LIMITE_REGISTROS).all()

    resultado = []
    for p, ticker, ativo_nome in rows:
        resultado.append({
            'id':             str(p.id),
            'data_pagamento': _fmt_data(p.data_pagamento),
            'data_com':       _fmt_data(p.data_com),
            'tipo':           p.tipo_provento.value if p.tipo_provento else '',
            'ticker':         ticker,
            'ativo':          ativo_nome,
            'quantidade':     _fmt_decimal_f(p.quantidade_ativos),
            'valor_por_acao': _fmt_decimal_f(p.valor_por_acao),
            'valor_bruto':    _fmt_decimal_f(p.valor_bruto),
            'imposto_retido': _fmt_decimal_f(p.imposto_retido),
            'valor_liquido':  _fmt_decimal_f(p.valor_liquido),
        })
    return resultado


def _query_posicoes(usuario_id: str, filtros: dict) -> list[dict]:
    from app.services.portfolio_service import PortfolioService

    metrics = PortfolioService.get_portfolio_metrics(usuario_id)
    posicoes_raw = metrics.get('posicoes', [])

    resultado = []
    for pos in posicoes_raw:
        if filtros.get('ativo_id') and str(pos.get('ativo_id', '')) != str(filtros['ativo_id']):
            continue
        resultado.append({
            'ticker':           pos.get('ticker', ''),
            'nome':             pos.get('nome', ''),
            'tipo':             pos.get('tipo', ''),
            'quantidade':       _fmt_decimal_f(pos.get('quantidade', 0)),
            'preco_medio':      _fmt_decimal_f(pos.get('preco_medio', 0)),
            'preco_atual':      _fmt_decimal_f(pos.get('preco_atual', 0)),
            'valor_investido':  _fmt_decimal_f(pos.get('valor_investido', 0)),
            'valor_atual':      _fmt_decimal_f(pos.get('valor_atual', 0)),
            'lucro_prejuizo':   _fmt_decimal_f(pos.get('lucro_prejuizo', 0)),
            'rentabilidade_pct': _fmt_decimal_f(pos.get('rentabilidade', 0)),
        })
    return resultado


# ---------------------------------------------------------------------------
# Renderers por formato
# ---------------------------------------------------------------------------

def _render_json(dados: list, meta: dict) -> tuple[bytes, str]:
    payload = {'meta': meta, 'dados': dados, 'total': len(dados)}
    content = json.dumps(payload, ensure_ascii=False, indent=2).encode('utf-8')
    return content, 'application/json'


def _render_csv(dados: list, meta: dict) -> tuple[bytes, str]:
    buf = io.StringIO()
    # Cabeçalho de metadados (sempre presente)
    buf.write(f"# Exitus — {meta.get('entidade', '').capitalize()}\n")
    buf.write(f"# Gerado em: {meta.get('gerado_em', '')}\n")
    buf.write(f"# Total de registros: {len(dados)}\n")
    buf.write(f"# Filtros: {json.dumps(meta.get('filtros', {}), ensure_ascii=False)}\n")

    if dados:
        writer = csv.DictWriter(buf, fieldnames=dados[0].keys(), delimiter=';')
        writer.writeheader()
        writer.writerows(dados)

    return buf.getvalue().encode('utf-8-sig'), 'text/csv; charset=utf-8'


def _render_excel(dados: list, meta: dict) -> tuple[bytes, str]:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = meta.get('entidade', 'dados').capitalize()

    # Linha de título
    ws['A1'] = f"Exitus — {meta.get('entidade', '').capitalize()}"
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = f"Gerado em: {meta.get('gerado_em', '')}  |  Total: {len(dados)} registros"
    ws['A2'].font = Font(italic=True, color='666666')
    ws.append([])  # linha em branco

    if not dados:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Cabeçalho de colunas
    headers = list(dados[0].keys())
    header_row = ws.max_row + 1
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h.replace('_', ' ').title())
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E79')
        cell.alignment = Alignment(horizontal='center')

    # Dados
    for row in dados:
        ws.append(list(row.values()))

    # Auto-ajuste de largura (aproximado)
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _render_pdf(dados: list, meta: dict) -> tuple[bytes, str]:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Título
    elements.append(Paragraph(
        f"<b>Exitus — {meta.get('entidade', '').capitalize()}</b>",
        styles['Title']
    ))
    elements.append(Paragraph(
        f"Gerado em: {meta.get('gerado_em', '')}  |  Total: {len(dados)} registros",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.5 * cm))

    if not dados:
        elements.append(Paragraph("Nenhum registro encontrado.", styles['Normal']))
        doc.build(elements)
        return buf.getvalue(), 'application/pdf'

    headers = list(dados[0].keys())
    col_labels = [h.replace('_', ' ').title() for h in headers]

    table_data = [col_labels]
    for row in dados:
        table_data.append([str(v) for v in row.values()])

    col_count = len(headers)
    page_width = landscape(A4)[0] - 3 * cm
    col_width = page_width / col_count

    table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0),  colors.HexColor('#1F4E79')),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  8),
        ('FONTSIZE',    (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF0FA')]),
        ('GRID',        (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',  (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)

    doc.build(elements)
    return buf.getvalue(), 'application/pdf'


# ---------------------------------------------------------------------------
# Entry point público
# ---------------------------------------------------------------------------

RENDERERS = {
    'json':  _render_json,
    'csv':   _render_csv,
    'excel': _render_excel,
    'pdf':   _render_pdf,
}

QUERIES = {
    'transacoes': _query_transacoes,
    'proventos':  _query_proventos,
    'posicoes':   _query_posicoes,
}

EXTENSOES = {
    'json':  'json',
    'csv':   'csv',
    'excel': 'xlsx',
    'pdf':   'pdf',
}


class ExportService:

    @staticmethod
    def exportar(usuario_id: str, entidade: str, formato: str, params: dict) -> tuple[bytes, str, str]:
        """
        Exporta dados de uma entidade no formato solicitado.

        Args:
            usuario_id: UUID do usuário
            entidade:   'transacoes' | 'proventos' | 'posicoes'
            formato:    'csv' | 'excel' | 'json' | 'pdf'
            params:     dict com filtros opcionais

        Returns:
            (conteúdo_bytes, content_type, filename)
        """
        if entidade not in QUERIES:
            raise BusinessRuleError(
                f"Entidade inválida: '{entidade}'. Use: {', '.join(QUERIES.keys())}."
            )
        if formato not in RENDERERS:
            raise BusinessRuleError(
                f"Formato inválido: '{formato}'. Use: {', '.join(RENDERERS.keys())}."
            )

        # Processar filtros de data
        filtros = {}
        if params.get('data_inicio'):
            filtros['data_inicio'] = _parse_date(params['data_inicio'], 'data_inicio')
        if params.get('data_fim'):
            filtros['data_fim'] = _parse_date(params['data_fim'], 'data_fim')
        if params.get('ativo_id'):
            filtros['ativo_id'] = params['ativo_id']
        if params.get('corretora_id'):
            filtros['corretora_id'] = params['corretora_id']
        if params.get('tipo'):
            filtros['tipo'] = params['tipo']

        # Buscar dados
        dados = QUERIES[entidade](usuario_id, filtros)

        # Metadados do cabeçalho
        meta = {
            'entidade':  entidade,
            'formato':   formato,
            'gerado_em': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'filtros':   {k: str(v) for k, v in filtros.items()},
            'usuario_id': str(usuario_id),
        }

        conteudo, content_type = RENDERERS[formato](dados, meta)

        ts = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'exitus_{entidade}_{ts}.{EXTENSOES[formato]}'

        return conteudo, content_type, filename
